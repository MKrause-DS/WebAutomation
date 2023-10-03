from time import sleep
from datetime import date, datetime
import os
import sys
import win32com.client as win32
from shutil import move
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import openpyxl

def write_to_protocol(status, message):
    protocol_file = f"Protocol_{date.today()}.xlsx"
    
    print(protocol_file)
    # Load the workbook or create a new one if it doesn't exist
    try:
        wb = openpyxl.load_workbook(protocol_file)
        ws = wb.active
        
        # If there's no active worksheet, create a new one
        if ws is None:
            ws = wb.create_sheet("Sheet1")
            
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['TimeStamp', 'Status', 'Message'])
    
    # Append the current date & time, status, and message to the Excel file
    ws.append([datetime.now().strftime('%Y-%m-%d %H:%M:%S'), status, message])

    wb.save(protocol_file)

def send_email(subject, recipient, body):
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)  # 0: olMailItem
    
    # Setup the parameters
    mail.Subject = subject
    mail.To = recipient
    mail.HTMLBody = body
    
    # Attach files
    mail.Attachments.Add(os.path.join(os.path.dirname(os.path.abspath(__file__)), f"Protocol_{date.today()}.xlsx"))
    directory = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Error_Screenshots")

    for filename in os.listdir(directory):
        if datetime.today().strftime('%Y-%m-%d') in filename:
            filepath = os.path.join(directory, filename)
            mail.Attachments.Add(filepath)
    
    # Send the email
    mail.Send()
    print("Email sent successfully.")

def initialization():
    # Current directory
    current_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Archive directory
    archive_dir = os.path.join(current_dir, "Archive")
    
    # Ensure the Archive directory exists
    if not os.path.exists(archive_dir):
        os.makedirs(archive_dir)

    # Today's date in the format YYYY-MM-DD
    today_str = date.today().strftime('%Y-%m-%d')
    
    # List all files in the directory
    for filename in os.listdir(current_dir):
        # Check for Protocol Excel files that don't have the current date in the filename
        if "Protocol" in filename and filename.endswith(".xlsx") and today_str not in filename:
            # Move the file to the Archive folder
            move(os.path.join(current_dir, filename), os.path.join(archive_dir, filename))

    # Create the webdriver
    driver = webdriver.Chrome('./chromedriver.exe')
    driver.get('https://neu.insolvenzbekanntmachungen.de/ap/suche.jsf')
    sleep(3)

    # Maximize the window
    driver.maximize_window()

    return driver


def search_company(CompanyName, driver):
    

    max_retries = 3
    error_counter = 0
    screenshot_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Error_Screenshots")
    

    while error_counter < max_retries:
        try:
            # Attempt to find the firmenname element
            firmenname = driver.find_element(By.ID, 'frm_suche:ireg_registereintrag:itx_registernummer')

            # If successful, send the company name
            firmenname.send_keys(CompanyName)
            sleep(10)
            write_to_protocol('Successfull','Company ' + CompanyName +'found')
            break
        except:
            error_counter += 1
            print(f"Attempt {error_counter} failed. Retrying...")
            write_to_protocol('Retry','Text')

            # Make Screenshot
            screenshot_path = os.path.join(screenshot_dir, f"Error_{CompanyName}_{datetime.today().strftime('%Y-%m-%d_%H %M %S')}.png")
            driver.save_screenshot(screenshot_path)

            # Close driver and reinitalize
            driver.quit()
            driver = initialization()

    if error_counter == max_retries:
        print("Failed to find the element after maximum retries. Exiting.")
        write_to_protocol('Failed','Text')

        # Make a screenshot
        screenshot_path = os.path.join(screenshot_dir, f"Error_{CompanyName}_{datetime.today().strftime('%Y-%m-%d_%H %M %S')}.png")
        driver.save_screenshot(screenshot_path)


        # Send Email after failed run
        send_email("Process Failed", "mario.krause@fau.de", "Hallo Herr Geßlein, <br> <br> irgendwas läuft hier übelst schief.")

        # Close the browser and end the process
        driver.quit()
        sys.exit(0)
        return 

if __name__ == "__main__":

    
    driver=initialization()

    search_company('asd',driver)
    sleep(5)
    driver.quit()

    send_email("Process sucessfull", "mario.krause@fau.de", "Hallo Herr Geßlein, <br> <br> es läuft alles nach Plan!")

